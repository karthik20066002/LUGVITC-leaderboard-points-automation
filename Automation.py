#-------------------------------------------------------------------------------
# Name:             Automation.py
# Purpose:          Automate the points table calculation for the LUGVITC
#
# Organization:     LUGVITC
#
# Created:          21 12 2024
# License:          Apache License Version 2.0
#
# Developed by:     Meit Sant [Github:MT_276]
#                   Karthik Ramanathan Lakshmanan [Github:karthik20066002] 
#-------------------------------------------------------------------------------
#> Notes to Meit:
#> Dude, everything was broken asf. Your code did not work at all.
#> I've rewritten almost everything in the program, except the reading of the data file and reading the rubrics.
#> Update 4:34AM, The Rubrics were also unusable. Reformatted.
#> If there is a '>' sign in front of the '#', the code has been changed by me, and you should really have a look at it.
#> Any issues, you know where to get me.
#> ~Axiss
#-------------------------------------------------------------------------------


print(f'Points table automation for LUGVITC')
print('Licence          : Apache License Version 2.0')
print('[PROCESS] Starting.')
import openpyxl, sys

# Checking if data file exits 
try:
    with open('data.txt', 'r') as file:
        data = file.readlines()
        path = data[1].split('"')[1]
        print(path)
        
        print("\n[INFO] Using path from data file")
except FileNotFoundError:
    # Taking the path from the user
    print("\n[INFO] Data file not found")
    path = input("[INPUT] Enter the path of the data file > ")
    path = path.replace('\\', '/')
    path = path.replace('"', '')
    
    # Creating a new data file
    with open('data.txt', 'w') as file:
        file.write(f'Path of the file:- \n > "{path}"\n\n')
except:
    print('Error occured while reading data file')
    exit()


# Get the file name from the path
print(f"[INFO] Chosen file : {path.split('/')[-1]}")


# Load the workbook
try:
    wb = openpyxl.load_workbook(path)
    print(f"[INFO] Workbook loaded successfully")
except:
    print(f"[ERROR] Error occured while loading the workbook")
    sys.exit()


#> Define all members list. This will be iterated through to add to the Excel sheet.
all_members_data=[]

# Load the rubrics sheet
sheet = wb['Rubric']


# Read the Rubrics data into a dictionary
rubrics = {}
for row in sheet.iter_rows(min_row=2, max_col=2, values_only=True):
    task, points = row
    task = task.split('(')
    task = task[0].strip()
    rubrics[task] = points

print(rubrics)

# Reading 'Technical Support Cyber-0-Day' sheet
sheet = wb['Tech Support C0D3']

#> Iterate through each row
for i in range(2,sheet.max_row+2):

    #> Define row data. Holds the directly scraped value, raw data only.
    row_data = []

    #> Define member dictionary. This dictionary will hold the properly formatted data.
    member_data = {
    'Reg No': '',
    'Name' : '',
    'Contributions' : '',
    'Events' : '',
    'Points' : 0
    }

    #> Grabbing each individual cell in the row
    for j in range(1,7): 
        cell_item = sheet.cell(row = i, column = j)
        row_data.append(cell_item.value)
    else:


    #> Calculating the points 
        if row_data[1]==None: #> We keep this to make sure no null rows clutter the sheet.
            pass


        else:
            member_data['Reg No'] = row_data[1] #> Adds the registration number to the dictionary
            member_data['Name'] = row_data[0] #> Adds the name to the dictionary

            if row_data[5] == 'Yes' or row_data[3] == 'Yes' or row_data[4] == 'Yes':
                member_data['Contributions'] += 'Tech Support, '
                member_data['Events'] += 'Cyber-0-Day 3.0, '
            if row_data[3] == 'Yes':
                member_data['Points'] += rubrics['Overnight']
            if row_data[4] == 'Yes':
                member_data['Points'] += rubrics['Technical Support']
            if row_data[5] == 'Yes':
                member_data['Points'] += rubrics['Technical Support']


            all_members_data.append(member_data) #> Appends it to the all member data list.

#> NOTE: The block below is temporarily commented due to it not searching for the occurances. Must fix.
# Searching reg no for other occurances in the workbook
# for sheet in wb.sheetnames:
#     print(sheet)
#     if sheet == 'Technical Support Cyber-0-Day':
#         continue
#     sheet = wb[sheet]
#     for row in sheet.iter_rows(min_row=2, values_only=True):
#         print(row)

# Close the workbook
wb.close()
print(r'[INFO] Data trawling complete.')


#> Saving the data to a new excel file. Opening new workbook.
print(r'[INFO] Opening new workbook.')
scraped_values = openpyxl.Workbook()
sheet_of_points = scraped_values.active #> New workbook handle


book_name = input(r"[INPUT] Enter workbook name >> ")
sheet_name = input(r"[INPUT] Enter worksheet name >> ")
sheet_of_points.title = r""+sheet_name

#> Grabbing the headings and writing them to the new excel sheet.
headers = list(member_data.keys())
for i in range(len(headers)):
    temp_cell = sheet_of_points.cell(row = 1, column = i+1)
    temp_cell.value = headers[i]

#> Iterating through the dictionary and writing the values
for j in range(1,len(all_members_data)):
    for i in range(len(headers)):
        value_cell = sheet_of_points.cell(row = j+1, column = i+1 )
        temp_cell = sheet_of_points.cell(row = 1, column = i+1)
        value_cell.value = all_members_data[j][temp_cell.value]
print('[INFO] Data written to buffer.')
#> And finally, save the excel sheet in .xlsx format.
#> Something to look at: How do I make the document format newer? "Excel 2007 spreadsheet" is kinda old.
print('[INFO] Data saved to Excel file in same directory.')
scraped_values.save(book_name+r".xlsx")
print('[INFO] Closing file handles.')
scraped_values.close()
print('[PROCESS] Complete. Terminating...')







